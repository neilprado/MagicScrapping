import openpyxl
from openpyxl.styles import PatternFill

def createExcelFile(three_items, three_prices, liga_items, liga_prices, currency, filename="sealed.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Three4One Item'
    sheet['B1'] = 'Three4One Price'
    sheet['C1'] = 'LigaMagic Item'
    sheet['D1'] = 'LigaMagic Price'
    sheet['E1'] = 'Convert Price from EUR to BRL'

    color_nice_to_buy = 'FFDCFDD9'
    color_must_buy = 'FF75BEFA'
    color_maybe_buy = 'FFFFBF46'
    color_must_not_buy = 'FF83344D'

    for item in range(len(three_items)):

        try:
            liga_price = float(liga_prices[item])
            three_price = float(three_prices[item])
        except ValueError:
            print(f"Erro ao converter liga_prices[{item}] para float.")
            continue

        sheet.cell(row=item + 2, column=1, value=three_items[item])
        sheet.cell(row=item + 2, column=2, value=three_prices[item])
        sheet.cell(row=item + 2, column=3, value=liga_items[item])
        sheet.cell(row=item + 2, column=4, value=liga_prices[item])
        sheet.cell(row=item + 2, column=5, value=three_price * currency)

        comparison_value = three_price * currency

        if comparison_value + 200 <= liga_price:
            color = color_must_buy
        elif comparison_value <= liga_price:
            color = color_nice_to_buy
        elif comparison_value + 200 >= liga_price:
            color = color_maybe_buy
        elif comparison_value >= liga_price:
            color = color_must_not_buy

        sheet.cell(row=item + 2, column=5).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    workbook.save(filename)