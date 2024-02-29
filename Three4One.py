import re

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
import openpyxl


def config():
    driver = webdriver.Safari()
    driver.maximize_window()
    driver.get("https://shop.threeforonetrading.com")

    return driver


def scrapping(searches):
    driver = config()

    items = []
    prices = []

    for search in searches:
        search_input = driver.find_element(By.CSS_SELECTOR, ".search-bar__input")
        search_input.clear()
        search_input.send_keys(search)

        search_input.send_keys(Keys.RETURN)

        time.sleep(5)

        product_items = driver.find_elements(By.CSS_SELECTOR, ".product-item")

        for item in product_items:
            title = item.find_element(By.CSS_SELECTOR, ".product-item__title").text
            if "MtG Advent Calendar" not in title and "Sol Ring" not in title:
                items.append(title)
                prices.append(extract_euro_price(item.find_element(By.CSS_SELECTOR, ".price").text))

    finishing(driver)

    return items, prices


def finishing(driver):
    driver.quit()


def extract_euro_price(price):
    pattern = re.compile(r'€\s*(\d+(?:\.\d+)?)')
    matching = pattern.search(price)

    if matching:
        euro_price = matching.group(1)
        return float(euro_price)
    else:
        return None


def createExcelFile(items, prices, filename="ThreeForOne.xlsx"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    sheet['A1'] = 'Item'
    sheet['B1'] = 'Price'

    for item in range(len(items)):
        sheet.cell(row=item + 2, column=1, value=items[item])
        sheet.cell(row=item + 2, column=2, value=prices[item])

    workbook.save(filename)
